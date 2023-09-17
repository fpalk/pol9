import pathlib
from io import StringIO
from sylk_parser import SylkParser
import pandas as pd
from win32com.client import Dispatch
from openpyxl import load_workbook
from openpyxl.chart.text import RichText
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.chart.reference import Reference
from openpyxl.chart.line_chart import LineChart
from openpyxl.drawing.text import RichTextProperties, Paragraph, ParagraphProperties, CharacterProperties

EMG_SYLK_DIR = "./data/emg"
EMG_CSV_DIR = "./data/emg_csv"
EMG_XLSX_DIR = "./data/emg_xlsx"
EMG_XLSX_WITH_CHARTS_DIR = "./data/emg_xlsx_with_charts"

def sylk_to_csv(src_path: str):

    for file_path in pathlib.Path(src_path).iterdir():
        parser = SylkParser(file_path)

        fbuf = StringIO()
        parser.to_csv(fbuf)

        results = fbuf.getvalue()

        with open(f"{src_path}_csv\\{file_path.stem}.csv", "w+", encoding='utf-8') as file:
            file.write(results)

        fbuf.close()

    print("ALL DONE👍👍👍👍👍👍👍👍")


def clean_emg(src_path: str):
    for file_path in pathlib.Path(src_path).iterdir():
        df = pd.read_csv(
            file_path, 
            header=3,
            index_col="Time,s",
            usecols=["Time,s", "FLEX.CARP.R,uV", "MED. GASTRO,uV"],
            dtype={"Time,s": float, "FLEX.CARP.R,uV": float, "MED. GASTRO,uV": float},
        )

        df.rename({"Time,s": "time(s)", "FLEX.CARP.R,uV": "근전도(팔,수근굴근)(uV)", "MED. GASTRO,uV": "근전도(종아리,비복근)(uV)"}, inplace=True)

        max_time = int(df.index.max())
        
        emg_indexer = df.index.get_indexer(
            [i * 0.1 for i in range(max_time * 10)], 
            method="nearest",
            limit=1,
            tolerance=0.05
        )

        df_new = df.iloc[emg_indexer, :]

        df_new.abs().to_excel(f"data/emg_xlsx/{file_path.stem}.xlsx")

        print("ALL DONE👍👍👍👍👍👍👍👍")

def plot_eeg(src_path: str):
    for file_path in pathlib.Path(src_path).iterdir():
        [subject_number, scenario_number] = file_path.stem.split("-")


        wb = load_workbook(file_path)
        ws = wb.active

        # ARM
        c1 = LineChart()
        c1.title = f"피실험자{subject_number} 시나리오{scenario_number} 근전도(팔,수근굴근)"
        c1.x_axis.title = "time(s)"
        c1.x_axis.tickLblSkip = 200
        c1.x_axis.tickLblPos = "low"
        c1.x_axis.txPr = RichText(bodyPr=RichTextProperties(anchor="ctr", anchorCtr="1", rot="-2700000",
                                                            spcFirstLastPara="1", vertOverflow="ellipsis", wrap="square"),
                                    p=[Paragraph(pPr=ParagraphProperties(
                                        defRPr=CharacterProperties()), endParaRPr=CharacterProperties())]
                                    )
        c1.y_axis.scaling.min = 0
        c1.y_axis.scaling.max = 250
        c1.y_axis.title = "근전도(팔,수근굴근)(uV)"
        c1.legend = None

        arm_data = Reference(ws, min_col=2, max_col=2,
                                 min_row=2, max_row=ws.max_row)

        label = Reference(ws, min_col=1, max_col=1,
                            min_row=2, max_row=ws.max_row)



        c1.add_data(arm_data)
        c1.set_categories(label)
        c1.series[0].graphicalProperties.line.width = pixels_to_EMU(1)
        c1.width = 20

        ws.add_chart(c1, f"{get_column_letter(5)}{2}")

        # LEG

        c2 = LineChart()
        c2.title = f"피실험자{subject_number} 시나리오{scenario_number} 근전도(종아리,비복근)"
        c2.x_axis.title = "time(s)"
        c2.x_axis.tickLblSkip = 200
        c2.x_axis.tickLblPos = "low"
        c2.x_axis.txPr = RichText(bodyPr=RichTextProperties(anchor="ctr", anchorCtr="1", rot="-2700000",
                                                            spcFirstLastPara="1", vertOverflow="ellipsis", wrap="square"),
                                    p=[Paragraph(pPr=ParagraphProperties(
                                        defRPr=CharacterProperties()), endParaRPr=CharacterProperties())]
                                    )
        c2.y_axis.scaling.min = 0
        c2.y_axis.scaling.max = 250
        c2.y_axis.title = "근전도(종아리,비복근)(uV)"


        c2.legend = None

        label = Reference(ws, min_col=1, max_col=1,
                            min_row=2, max_row=ws.max_row)
        
        leg_data = Reference(ws, min_col=3, max_col=3,
                                 min_row=2, max_row=ws.max_row)
        
        c2.add_data(leg_data)
        c2.set_categories(label)
        c2.series[0].graphicalProperties.line.width = pixels_to_EMU(1)
        c2.width = 20



        ws.add_chart(c2, f"{get_column_letter(5)}{15}")


        wb.save(f"{EMG_XLSX_WITH_CHARTS_DIR}/{file_path.stem}.xlsx")

    print("ALL DONE👍👍👍👍👍👍👍👍")
        


def export_image():
    app = Dispatch("Excel.Application")
    # It's important to use the absolute path, it won't work with a relative one.
    workbook = app.Workbooks.Open(Filename=workbook_file_name)

    app.DisplayAlerts = False

    for i, sheet in enumerate(workbook.Worksheets):
        for chartObject in sheet.ChartObjects():
            print(sheet.Name + ':' + chartObject.Name)
            # It's important to use the absolute path, it won't work with a relative one.
            chartObject.Chart.Export(str(pathlib.Path().resolve()) + "\chart" + str(i+1) + ".png")

    workbook.Close(SaveChanges=False, Filename=workbook_file_name)


# sylk_to_csv(emg_sylk_src_path)
# clean_emg(EMG_CSV_DIR)

plot_eeg(EMG_XLSX_DIR)